import pandas as pd
import openpyxl
from typing import List

def excel_sheet_list(inputfile: str) -> List[str]:
    inputfile = "ch4_project_characterization.xlsx"
    data_sheet= []
    workbook = openpyxl.load_workbook(inputfile)
    sheetnames = workbook.sheetnames
    for i in list(range(len(sheetnames))):
        if sheetnames[i].find(" - ") > 0:
            data_sheet.append(sheetnames[i])
    return data_sheet

def list_to_lower(data_sheet: List[str]) -> List[str]:
    data_sheet_lower =[]
    for i in list(range(len(data_sheet))):
        data_sheet_lower.append(data_sheet[i].lower().replace(" ", "").replace("-", "_"))
    return data_sheet_lower

def col_to_lists(inputfile: str) -> List[str]:
    data_sheet = excel_sheet_list(inputfile)
    data_sheet_lower = list_to_lower(data_sheet)
    result= []
    for i in list(range(len(data_sheet))):
        df = pd.read_excel(inputfile, sheet_name = data_sheet[i])
        df = df.drop(columns = 'Month (from time 0)')
        oil_list = df["Gross Oil (bbl)"].astype(float).to_list()
        gas_list = df["Gross Gas (mcf)"].astype(float).to_list()
        capex_list = df["Total Capital (M$)"].astype(float).to_list()
        opex_list = df["Total Operating Expense (M$)"].astype(float).to_list()
        result.append((f'{data_sheet_lower[i]}_oil_list:{oil_list}'))
        result.append((f'{data_sheet_lower[i]}_gas_list:{gas_list}'))
        result.append((f'{data_sheet_lower[i]}_capex_list:{capex_list}'))
        result.append((f'{data_sheet_lower[i]}_opex_list:{opex_list}'))
    return result 

def list_filter(name: str, list: List[str]) -> List[str]:
    filtered_list = [item for item in list if 'nmb' in item]
    return filtered_list

